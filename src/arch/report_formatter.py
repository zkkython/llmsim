"""
性能报告格式化输出 - 支持多种输出格式（Console, Excel等）
"""
from abc import ABC, abstractmethod
from typing import List, Dict, Any
from pathlib import Path
from src.arch.perf_calculator import ModelPerformance


class ReportFormatter(ABC):
    """性能报告格式化器的抽象基类"""
    
    @abstractmethod
    def format(self, model_perf: ModelPerformance) -> Any:
        """
        格式化性能报告
        
        Args:
            model_perf: 模型性能指标
            
        Returns:
            格式化后的输出（内容取决于具体实现）
        """
        pass
    
    @abstractmethod
    def save(self, model_perf: ModelPerformance, output_path: str = None) -> None:
        """
        保存性能报告到文件或输出
        
        Args:
            model_perf: 模型性能指标
            output_path: 输出文件路径（可选）
        """
        pass
    
    def _collect_data(self, model_perf: ModelPerformance) -> List[Dict[str, Any]]:
        """
        收集所有性能数据供格式化器使用
        
        Args:
            model_perf: 模型性能指标
            
        Returns:
            包含所有行数据的列表
        """
        all_rows = []
        for layer_perf in model_perf.layer_performances:
            for op_perf in layer_perf.operators:
                op_meta = op_perf.metadata
                time_us = op_perf.total_time * 1000.0
                percentage = model_perf.get_percentage(time_us)
                all_rows.append({
                    'name': op_perf.name,
                    'type': op_perf.op_type,
                    'm': op_meta.io_config.input_shape.m,
                    'n': op_meta.io_config.output_shape.n,
                    'k': op_meta.io_config.input_shape.n,
                    'batch': op_meta.batch_size,
                    'layers': op_meta.num_layers,
                    'in_dtype': op_meta.io_config.input_dtype.name,
                    'out_dtype': op_meta.io_config.output_dtype.name,
                    'weight_dtype': op_meta.io_config.weight_dtype.name,
                    'compute': op_perf.compute_time,
                    'memory': op_perf.memory_time,
                    'transfer': op_perf.transfer_time,
                    'total': op_perf.total_time,
                    'percent': percentage,
                })
        return all_rows


class ConsoleReportFormatter(ReportFormatter):
    """控制台输出格式化器 - 美化的表格格式"""
    
    @staticmethod
    def _display_width(text: str) -> int:
        """计算文本的显示宽度（考虑中文字符占2个宽度）"""
        width = 0
        for char in text:
            if ord(char) >= 0x4E00 and ord(char) <= 0x9FFF:  # 中文字符范围
                width += 2
            else:
                width += 1
        return width
    
    @staticmethod
    def _pad_string(text: str, width: int, align: str = 'left') -> str:
        """根据显示宽度填充字符串"""
        display_len = ConsoleReportFormatter._display_width(text)
        padding = width - display_len
        if padding <= 0:
            return text
        if align == 'left':
            return text + ' ' * padding
        else:
            return ' ' * padding + text
    
    def format(self, model_perf: ModelPerformance) -> None:
        """
        格式化并打印性能报告到控制台
        
        Args:
            model_perf: 模型性能指标
        """
        all_rows = self._collect_data(model_perf)
        
        # 计算每列的最大宽度（考虑中文字符）
        if all_rows:
            col_widths = {}
            # 文本列
            col_widths['name'] = max(
                max(self._display_width(row['name']) for row in all_rows),
                self._display_width('算子名称')
            ) + 2
            col_widths['type'] = max(
                max(self._display_width(row['type']) for row in all_rows),
                self._display_width('类型')
            ) + 1
            # 数值列
            col_widths['m'] = max(max(len(str(row['m'])) for row in all_rows), len('m')) + 1
            col_widths['n'] = max(max(len(str(row['n'])) for row in all_rows), len('n')) + 1
            col_widths['k'] = max(max(len(str(row['k'])) for row in all_rows), len('k')) + 1
            col_widths['batch'] = max(max(len(str(row['batch'])) for row in all_rows), len('batch')) + 1
            col_widths['layers'] = max(max(len(str(row['layers'])) for row in all_rows), len('layers')) + 1
            col_widths['in_dtype'] = max(
                max(self._display_width(row['in_dtype']) for row in all_rows),
                self._display_width('输入')
            ) + 1
            col_widths['out_dtype'] = max(
                max(self._display_width(row['out_dtype']) for row in all_rows),
                self._display_width('输出')
            ) + 1
            col_widths['weight_dtype'] = max(
                max(self._display_width(row['weight_dtype']) for row in all_rows),
                self._display_width('权重')
            ) + 1
            # 浮点数列（宽度足以显示格式化后的值）
            col_widths['compute'] = max(len('计算(us)'), 10) + 1
            col_widths['memory'] = max(len('内存(us)'), 10) + 1
            col_widths['transfer'] = max(len('传输(us)'), 10) + 1
            col_widths['total'] = max(len('总时间(ms)'), 10) + 1
            col_widths['percent'] = max(len('占比(%)'), 8) + 1
        else:
            col_widths = {
                'name': 20, 'type': 10, 'm': 6, 'n': 6, 'k': 6,
                'batch': 8, 'layers': 8, 'in_dtype': 8, 'out_dtype': 8,
                'weight_dtype': 8, 'compute': 12, 'memory': 12, 'transfer': 12,
                'total': 12, 'percent': 10
            }
        
        # 计算总宽度（考虑分隔符）
        total_width = sum(col_widths.values()) + 16  # 16 = 分隔符号的宽度
        
        # 打印表头
        print()
        print("┌" + "─" * (total_width - 2) + "┐")
        header_text = f"性能分析报告: {model_perf.model_name} ({model_perf.forward_mode})"
        padding = total_width - self._display_width(header_text) - 4
        print(f"│ {header_text}" + " " * max(padding, 1) + " │")
        print("├" + "─" * (total_width - 2) + "┤")
        
        # 打印列标题
        headers = [
            ('name', '算子名称'),
            ('type', '类型'),
            ('m', 'm'),
            ('n', 'n'),
            ('k', 'k'),
            ('batch', 'batch'),
            ('layers', 'layers'),
            ('in_dtype', '输入'),
            ('out_dtype', '输出'),
            ('weight_dtype', '权重'),
            ('compute', '计算(us)'),
            ('memory', '内存(us)'),
            ('transfer', '传输(us)'),
            ('total', '总时间(ms)'),
            ('percent', '占比(%)'),
        ]
        
        header_line = "│ "
        for key, label in headers:
            width = col_widths[key]
            if key in ['compute', 'memory', 'transfer', 'total', 'percent']:
                # 数值列右对齐
                padded = self._pad_string(label, width, 'right')
            else:
                # 文本列左对齐
                padded = self._pad_string(label, width, 'left')
            header_line += f"{padded} │ "
        print(header_line)
        print("├" + "─" * (total_width - 2) + "┤")
        
        # 打印数据行
        for row in all_rows:
            line = "│ "
            for key, _ in headers:
                width = col_widths[key]
                value = row[key]
                
                if key in ['compute', 'memory', 'transfer']:
                    # 浮点数，3位小数
                    formatted = f"{value:.3f}"
                    padded = self._pad_string(formatted, width, 'right')
                    line += f"{padded} │ "
                elif key == 'total':
                    # 总时间，3位小数
                    formatted = f"{value:.3f}"
                    padded = self._pad_string(formatted, width, 'right')
                    line += f"{padded} │ "
                elif key == 'percent':
                    # 百分比，2位小数
                    formatted = f"{value:.2f}"
                    padded = self._pad_string(formatted, width, 'right')
                    line += f"{padded} │ "
                elif key in ['m', 'n', 'k', 'batch', 'layers']:
                    # 整数列
                    formatted = str(value)
                    padded = self._pad_string(formatted, width, 'right')
                    line += f"{padded} │ "
                else:
                    # 文本列
                    padded = self._pad_string(str(value), width, 'left')
                    line += f"{padded} │ "
            print(line)
        
        print("├" + "─" * (total_width - 2) + "┤")
        
        # 汇总统计
        total_compute_ms = model_perf.total_compute_time / 1000.0
        total_memory_ms = model_perf.total_memory_time / 1000.0
        total_transfer_ms = model_perf.total_transfer_time / 1000.0
        total_ms = model_perf.total_time / 1000.0
        
        summary_lines = [
            f"总结统计 / 单层",
            f"  计算时间: {total_compute_ms:>10.3f} ms",
            f"  内存时间: {total_memory_ms:>10.3f} ms",
            f"  传输时间: {total_transfer_ms:>10.3f} ms",
            f"  总耗时:   {total_ms:>10.3f} ms",
        ]
        
        for summary in summary_lines:
            padding = total_width - self._display_width(summary) - 4
            print(f"│ {summary}" + " " * max(padding, 1) + " │")
        
        # 性能瓶颈
        bottleneck = model_perf.get_bottleneck_op()
        if bottleneck:
            layer_name, op_name, op_perf = bottleneck
            print("├" + "─" * (total_width - 2) + "┤")
            bottleneck_text = f"性能瓶颈: {op_name} (总耗时: {op_perf.total_time:.3f} ms)"
            padding = total_width - self._display_width(bottleneck_text) - 4
            print(f"│ {bottleneck_text}" + " " * max(padding, 1) + " │")
        
        print("└" + "─" * (total_width - 2) + "┘")
    
    def save(self, model_perf: ModelPerformance, output_path: str = None) -> None:
        """
        保存性能报告到文件
        
        Args:
            model_perf: 模型性能指标
            output_path: 输出文件路径（可选）
        """
        if output_path:
            # 重定向标准输出到文件
            import sys
            from io import StringIO
            
            old_stdout = sys.stdout
            sys.stdout = StringIO()
            
            self.format(model_perf)
            
            output = sys.stdout.getvalue()
            sys.stdout = old_stdout
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(output)
            print(f"报告已保存到: {output_path}")
        else:
            # 直接打印到控制台
            self.format(model_perf)


class ExcelReportFormatter(ReportFormatter):
    """Excel输出格式化器"""
    
    def format(self, model_perf: ModelPerformance) -> Any:
        """
        格式化性能报告为Excel工作簿
        
        Args:
            model_perf: 模型性能指标
            
        Returns:
            openpyxl.Workbook 对象
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "需要安装 openpyxl 库来支持Excel输出。"
                "请运行: pip install openpyxl"
            )
        
        all_rows = self._collect_data(model_perf)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "性能分析"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 10
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 添加标题
        ws['A1'] = f"性能分析报告: {model_perf.model_name} ({model_perf.forward_mode})"
        ws['A1'].font = title_font
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = center_align
        
        # 添加列标题
        headers = [
            '算子名称', '类型', 'm', 'n', 'k', 'batch', 'layers',
            '输入', '输出', '权重', '计算(us)', '内存(us)', '传输(us)',
            '总时间(ms)', '占比(%)'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # 添加数据行
        col_keys = [
            'name', 'type', 'm', 'n', 'k', 'batch', 'layers',
            'in_dtype', 'out_dtype', 'weight_dtype', 'compute', 'memory',
            'transfer', 'total', 'percent'
        ]
        
        for row_idx, row_data in enumerate(all_rows, start=4):
            for col_idx, key in enumerate(col_keys, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data[key]
                
                # 格式化值
                if key in ['compute', 'memory', 'transfer', 'total']:
                    cell.value = round(value, 3)
                    cell.number_format = '0.000'
                    cell.alignment = right_align
                elif key == 'percent':
                    cell.value = round(value, 2)
                    cell.number_format = '0.00'
                    cell.alignment = right_align
                elif key in ['m', 'n', 'k', 'batch', 'layers']:
                    cell.value = value
                    cell.alignment = right_align
                else:
                    cell.value = str(value)
                    cell.alignment = left_align
                
                cell.border = border
        
        # 添加统计行
        stats_row = len(all_rows) + 5
        total_compute_ms = model_perf.total_compute_time / 1000.0
        total_memory_ms = model_perf.total_memory_time / 1000.0
        total_transfer_ms = model_perf.total_transfer_time / 1000.0
        total_ms = model_perf.total_time / 1000.0
        
        stats_data = [
            ('计算时间 (ms)', total_compute_ms),
            ('内存时间 (ms)', total_memory_ms),
            ('传输时间 (ms)', total_transfer_ms),
            ('总耗时 (ms)', total_ms),
        ]
        
        for idx, (label, value) in enumerate(stats_data):
            label_cell = ws.cell(row=stats_row + idx, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            
            value_cell = ws.cell(row=stats_row + idx, column=2)
            value_cell.value = round(value, 3)
            value_cell.number_format = '0.000'
        
        # 添加性能瓶颈信息
        bottleneck = model_perf.get_bottleneck_op()
        if bottleneck:
            bottleneck_row = stats_row + len(stats_data) + 2
            label_cell = ws.cell(row=bottleneck_row, column=1)
            label_cell.value = "性能瓶颈"
            label_cell.font = Font(bold=True)
            
            _, op_name, op_perf = bottleneck
            value_cell = ws.cell(row=bottleneck_row, column=2)
            value_cell.value = f"{op_name} (总耗时: {op_perf.total_time:.3f} ms)"
        
        return wb
    
    def save(self, model_perf: ModelPerformance, output_path: str = None) -> None:
        """
        保存性能报告到Excel文件
        
        Args:
            model_perf: 模型性能指标
            output_path: 输出Excel文件路径
        """
        if not output_path:
            output_path = "性能报告.xlsx"
        
        try:
            wb = self.format(model_perf)
            wb.save(output_path)
            print(f"Excel报告已保存到: {output_path}")
        except Exception as e:
            print(f"保存Excel报告失败: {e}")


def create_formatter(format_type: str = 'console') -> ReportFormatter:
    """
    工厂函数 - 创建对应格式的报告格式化器
    
    Args:
        format_type: 输出格式类型 ('console' 或 'excel')
        
    Returns:
        对应的ReportFormatter实例
        
    Raises:
        ValueError: 如果格式类型不支持
    """
    if format_type.lower() == 'console':
        return ConsoleReportFormatter()
    elif format_type.lower() == 'excel':
        return ExcelReportFormatter()
    else:
        raise ValueError(f"不支持的输出格式: {format_type}。支持的格式: console, excel")
