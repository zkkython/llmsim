from src.arch.perf_calculator import ModelPerformance
from src.visual.report_base import ReportFormatter
from src.arch.config import ForwardMode
from typing import Any

class ExcelReportFormatter(ReportFormatter):
    """Excel输出格式化器"""
    
    def format(self, model_perf: ModelPerformance) -> Any:
        """
        格式化性能报告为Excel工作簿
        
        Args:
            model_perf: 模型性能指标
            
        Returns:
            openpyxl.Workbook 对象
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "需要安装 openpyxl 库来支持Excel输出。"
                "请运行: pip install openpyxl"
            )
        
        all_rows = self._collect_data(model_perf)
        
        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "性能分析"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 10
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 添加标题
        ws['A1'] = f"性能分析报告: {model_perf.model_name} ({model_perf.forward_mode})"
        ws['A1'].font = title_font
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = center_align
        
        # 添加列标题
        headers = [
            '算子名称', '类型', 'm', 'n', 'k', 'batch', 'layers',
            '输入', '输出', '权重', '计算(us)', '内存(us)', '传输(us)','单层理论延时(us)',
            '总时间(ms)', '占比(%)'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # 添加数据行
        col_keys = [
            'name', 'type', 'm', 'n', 'k', 'batch', 'layers',
            'in_dtype', 'out_dtype', 'weight_dtype', 'compute', 'memory',
            'transfer', 'op_time_single_layer', 'total', 'percent'
        ]
        
        for row_idx, row_data in enumerate(all_rows, start=4):
            for col_idx, key in enumerate(col_keys, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data[key]
                
                # 格式化值
                if key in ['compute', 'memory', 'transfer', 'total']:
                    cell.value = round(value, 3)
                    cell.number_format = '0.000'
                    cell.alignment = right_align
                elif key == 'op_time_single_layer':
                    cell.value = round(value, 3)
                    cell.number_format = '0.000'
                    cell.alignment = right_align
                elif key == 'percent':
                    cell.value = round(value, 2)
                    cell.number_format = '0.00'
                    cell.alignment = right_align
                elif key in ['m', 'n', 'k', 'batch', 'layers']:
                    cell.value = value
                    cell.alignment = right_align
                else:
                    cell.value = str(value)
                    cell.alignment = left_align
                
                cell.border = border
        
        # 添加统计行
        stats_row = len(all_rows) + 7
        total_compute_ms = model_perf.total_compute_time / 1000.0
        total_memory_ms = model_perf.total_memory_time / 1000.0
        total_transfer_ms = model_perf.total_transfer_time / 1000.0
        total_ms = model_perf.total_time / 1000.0
        
        stats_data = [
            ('计算时间 (ms)', total_compute_ms),
            ('内存时间 (ms)', total_memory_ms),
            ('传输时间 (ms)', total_transfer_ms),
            ('总耗时 (ms)', total_ms),
        ]
        
        for idx, (label, value) in enumerate(stats_data):
            label_cell = ws.cell(row=stats_row + idx, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
            
            value_cell = ws.cell(row=stats_row + idx, column=2)
            value_cell.value = round(value, 3)
            value_cell.number_format = '0.000'
        
        # 添加性能瓶颈信息
        bottleneck = model_perf.get_bottleneck_op()
        if bottleneck:
            bottleneck_row = stats_row + len(stats_data) + 2
            label_cell = ws.cell(row=bottleneck_row, column=1)
            label_cell.value = "性能瓶颈"
            label_cell.font = Font(bold=True)
                    
            _, op_name, op_perf = bottleneck
            value_cell = ws.cell(row=bottleneck_row, column=2)
            value_cell.value = f"{op_name} (总耗时: {op_perf.total_time:.3f} ms)"
                
        # TTFT 与 Throughput
        ttft = model_perf.get_ttft_or_tpot()
        throughput = model_perf.get_throughput()
        other_data = [
            ('TTFT (ms)', ttft) if model_perf.forward_mode == "EXTEND" else ('TPOT (ms)', ttft),
            ('吞吐量TPS', throughput),
        ]
        for idx, (label, value) in enumerate(other_data):
            ttft_throughput_row = stats_row + len(stats_data) + 2 + (2 if bottleneck else 0)
            label_cell = ws.cell(row=ttft_throughput_row + idx, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)
                    
            value_cell = ws.cell(row=ttft_throughput_row + idx, column=2)
            value_cell.value = round(value, 3)
            value_cell.number_format = '0.000'
                
        return wb
    
    def save(self, model_perf: ModelPerformance, output_path: str = None) -> None:
        """
        保存性能报告到Excel文件
        
        Args:
            model_perf: 模型性能指标
            output_path: 输出Excel文件路径
        """
        if not output_path:
            output_path = "性能报告.xlsx"
        
        try:
            wb = self.format(model_perf)
            wb.save(output_path)
            print(f"Excel报告已保存到: {output_path}")
        except Exception as e:
            print(f"保存Excel报告失败: {e}")
